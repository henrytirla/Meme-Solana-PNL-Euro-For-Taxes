import sqlite3
from dotenv import load_dotenv
import os
import requests
import pandas as pd
from dune_client.client import DuneClient
from dune_client.query import QueryBase
from dune_client.types import QueryParameter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


class SOLReport:
    def __init__(self, wallet_address, days_back=15):
        self.wallet_address = wallet_address
        self.validate_wallet_address()
        self.days_back = days_back
        load_dotenv()
        self.dune_api_key = os.getenv('DUNE_API_KEY')
        self.request_timeout = int(os.getenv('DUNE_API_REQUEST_TIMEOUT'))
        self.dune = DuneClient(
            api_key=self.dune_api_key,
            base_url="https://api.dune.com",
            request_timeout=self.request_timeout
        )
        self.TRANSACTION_QUERY_ID = 5572790

        self.parameters = [
            QueryParameter.text_type(name='day', value=f'-{self.days_back}'),
            QueryParameter.text_type(name='wallet', value=self.wallet_address)
        ]
        self.SOL_TRANSFER_QUERY_ID = 5585395
        #
        self.parameters_transfer = [QueryParameter.text_type(name='day', value=f'-{self.days_back}'),
                                 QueryParameter.text_type(name='Wallet', value=self.wallet_address)
                                  ]

        self.reports_folder = "final_reports"
        os.makedirs(self.reports_folder, exist_ok=True)
        self.output_file_path = os.path.join(self.reports_folder, f"{self.wallet_address}.xlsx")

        self.transaction_df = None
        self.sol_transfers_df= None
        self.solana_eur_price = self.get_sol_price_eur()
        self.db_name = "final.db"

        # Initialize database connection and create tables
        self.conn = sqlite3.connect(self.db_name)
        self.create_tables()
        self.wallet_id = self.get_or_create_wallet()

    def create_tables(self):
        """Create database tables if they don't exist"""
        cursor = self.conn.cursor()

        # Create wallets table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS wallets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                wallet_address TEXT UNIQUE NOT NULL,
                wallet_name TEXT,
                description TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')


        cursor.execute('''
             CREATE TABLE IF NOT EXISTS wallet_transactions (
                 wallet_id INTEGER NOT NULL,
                 token_symbol TEXT,
                 time_traded TEXT,
                 incoming REAL,
                 outcome REAL,
                 delta_token REAL,
                 spent_amount REAL,
                 earned_amount REAL,
                 spent_amount_eur REAL,
                 earned_amount_eur REAL,
                 number_buys INTEGER,
                 number_sells INTEGER,
                 delta_sol REAL,
                 delta_percentage REAL,
                 dexscreener TEXT,
                 block_time TEXT,
                 sol_eur_price REAL,
                 created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                 FOREIGN KEY (wallet_id) REFERENCES wallets (id)
             )
         ''')

        # Create index for better query performance
        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_wallet_transactions_wallet_id 
            ON wallet_transactions(wallet_id)
        ''')

        cursor.execute('''
            CREATE INDEX IF NOT EXISTS idx_wallet_transactions_block_time 
            ON wallet_transactions(block_time)
        ''')

        # Add SOL transfers table
        cursor.execute('''
              CREATE TABLE IF NOT EXISTS sol_transfers (
                  wallet_id INTEGER NOT NULL,
                  sol_eur_price REAL,
                  block_month TEXT,
                  from_owner TEXT,
                  to_owner TEXT,
                  sol_amount REAL,
                  sol_amount_eur REAL,
                  transaction_label TEXT,
                  solscan_link TEXT,
                  created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  FOREIGN KEY (wallet_id) REFERENCES wallets (id)
              )
          ''')

        # Create index for SOL transfers
        cursor.execute('''
              CREATE INDEX IF NOT EXISTS idx_sol_transfers_wallet_id 
              ON sol_transfers(wallet_id)
          ''')

        self.conn.commit()

    def get_or_create_wallet(self):
        """Get existing wallet ID or create new wallet record"""
        cursor = self.conn.cursor()

        # Check if wallet exists
        cursor.execute('SELECT id FROM wallets WHERE wallet_address = ?', (self.wallet_address,))
        result = cursor.fetchone()

        if result:
            wallet_id = result[0]
            print(f"Found existing wallet with ID: {wallet_id}")

            # Update last accessed time
            cursor.execute('''
                UPDATE wallets SET updated_at = CURRENT_TIMESTAMP 
                WHERE id = ?
            ''', (wallet_id,))

        else:
            # Create new wallet
            cursor.execute('''
                INSERT INTO wallets (wallet_address, wallet_name) 
                VALUES (?, ?)
            ''', (self.wallet_address, f"Wallet_{self.wallet_address[:8]}..."))

            wallet_id = cursor.lastrowid
            print(f"Created new wallet with ID: {wallet_id}")

        self.conn.commit()
        return wallet_id

    def get_sol_price_eur(self):
        try:
            url = "https://api.coingecko.com/api/v3/simple/price"
            params = {
                "ids": "solana",
                "vs_currencies": "eur"
            }
            response = requests.get(url, params=params)
            response.raise_for_status()
            price = response.json()['solana']['eur']
            print(f"Current SOL/EUR price: €{price}")
            return price
        except Exception as e:
            print(f"Error fetching SOL price: {e}")
            return 0  # Default fallback

    def validate_wallet_address(self):
        """Validate Solana wallet address format"""
        if not self.wallet_address or len(self.wallet_address) < 32:
            raise ValueError("Invalid Solana wallet address")
        return True

    """----------APPLY FORMATTING TO EXCEL REPORTS-----------------------------------------------------"""

    def combine_and_format_sheets_integrated(self, input_file_path=None, output_file_path=None):
        """Combine Summary and Transactions sheets into one formatted sheet - integrated version"""

        # Use the class output file path if none provided
        if input_file_path is None:
            input_file_path = self.output_file_path

        # Use same file if no output specified
        if output_file_path is None:
            output_file_path = input_file_path

        # Load the existing workbook
        workbook = load_workbook(input_file_path)

        # Check if required sheets exist
        if 'Summary' not in workbook.sheetnames or 'Transactions' not in workbook.sheetnames:
            print("❌ Required 'Summary' and 'Transactions' sheets not found. Skipping formatting.")
            return workbook

        # Access existing sheets
        summary_sheet = workbook['Summary']
        transactions_sheet = workbook['Transactions']

        # Create a new combined sheet
        if 'Summary and Transactions' in workbook.sheetnames:
            del workbook['Summary and Transactions']  # Remove if exists

        combined_sheet = workbook.create_sheet('Summary and Transactions', 0)  # Insert at beginning

        # Copy summary data (headers + data) with sol_price insertion
        current_row = 1

        # Find wallet_id column position
        wallet_id_col = None
        for col in range(1, summary_sheet.max_column + 1):
            header_value = summary_sheet.cell(row=1, column=col).value
            if header_value and 'wallet_id' in str(header_value).lower():
                wallet_id_col = col
                break

        # Copy summary headers with sol_price insertion
        new_col = 1
        for col in range(1, summary_sheet.max_column + 1):
            # Copy the current column
            summary_value = summary_sheet.cell(row=1, column=col).value
            combined_sheet.cell(row=current_row, column=new_col).value = summary_value
            new_col += 1

            # If this was the wallet_id column, insert sol_price_eur next
            if col == wallet_id_col:
                combined_sheet.cell(row=current_row, column=new_col).value = 'sol_price_eur'
                new_col += 1

        current_row += 1

        # Copy summary data with sol_price insertion
        for row in range(2, summary_sheet.max_row + 1):
            new_col = 1
            for col in range(1, summary_sheet.max_column + 1):
                # Copy the current column data
                summary_value = summary_sheet.cell(row=row, column=col).value
                combined_sheet.cell(row=current_row, column=new_col).value = summary_value
                new_col += 1

                # If this was the wallet_id column, insert sol_price value next
                if col == wallet_id_col:
                    combined_sheet.cell(row=current_row,
                                        column=new_col).value = self.solana_eur_price  # Use actual price
                    new_col += 1

            current_row += 1

        # Add empty row for spacing
        current_row += 1

        # Get columns to exclude from transactions
        exclude_columns = []
        for col in range(1, transactions_sheet.max_column + 1):
            header_value = transactions_sheet.cell(row=1, column=col).value
            if header_value:
                header_str = str(header_value).lower()
                if 'wallet_id' in header_str or 'sol_eur_price' in header_str or 'sol_price' in header_str:
                    exclude_columns.append(col)

        # Copy transaction headers (excluding unwanted columns)
        transaction_start_row = current_row
        new_col = 1
        for col in range(1, transactions_sheet.max_column + 1):
            if col not in exclude_columns:
                transaction_value = transactions_sheet.cell(row=1, column=col).value
                combined_sheet.cell(row=current_row, column=new_col).value = transaction_value
                new_col += 1

        current_row += 1

        # Copy transaction data (excluding unwanted columns)
        for row in range(2, transactions_sheet.max_row + 1):
            new_col = 1
            for col in range(1, transactions_sheet.max_column + 1):
                if col not in exclude_columns:
                    transaction_value = transactions_sheet.cell(row=row, column=col).value
                    combined_sheet.cell(row=current_row, column=new_col).value = transaction_value
                    new_col += 1
            current_row += 1

        # Apply formatting to the combined sheet
        self.apply_combined_formatting_integrated(combined_sheet, transaction_start_row)

        # Remove original sheets
        del workbook['Summary']
        del workbook['Transactions']

        # Save the workbook
        workbook.save(output_file_path)
        print(f'✅ Combined and formatted sheet saved to {output_file_path}')

        return workbook

    def apply_combined_formatting_integrated(self, worksheet, transaction_start_row):
        """Apply formatting to the combined sheet - integrated version"""

        # Define color fills
        brown_fill = PatternFill(start_color="A52A2A", end_color="A52A2A", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

        # Apply basic formatting to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Bold headers (summary header = row 1, transaction header = transaction_start_row)
                if cell.row == 1 or cell.row == transaction_start_row:
                    cell.font = Font(bold=True)

        # Adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        # Find summary columns (row 1)
        summary_cols = {}
        for cell in worksheet[1]:  # Row 1
            if cell.value:
                col_name = str(cell.value).lower()
                if 'total_spent_amount' in col_name:
                    summary_cols['total_spent_amount'] = cell.column
                elif 'actual_profit' in col_name:
                    summary_cols['actual_profit'] = cell.column
                elif 'pnl_realized_profits' in col_name:
                    summary_cols['pnl_r'] = cell.column
                elif 'pnl_realized_losses' in col_name:
                    summary_cols['pnl_l'] = cell.column

        # Find transaction columns (transaction_start_row)
        transaction_cols = {}
        for cell in worksheet[transaction_start_row]:
            if cell.value:
                col_name = str(cell.value).lower()
                if 'delta_sol' in col_name:
                    transaction_cols['delta_sol'] = cell.column
                elif 'delta_percentage' in col_name:
                    transaction_cols['delta_percentage'] = cell.column
                elif 'dexscreener' in col_name:
                    transaction_cols['dexscreener'] = cell.column
                elif 'number_buys' in col_name:
                    transaction_cols['number_buys'] = cell.column
                elif 'outcome' in col_name:
                    transaction_cols['outcome'] = cell.column
                elif 'incoming' in col_name:
                    transaction_cols['incoming'] = cell.column

        # Format summary data (row 2)
        if 'total_spent_amount' in summary_cols and 'pnl_r' in summary_cols:
            total_spent_cell = worksheet.cell(row=2, column=summary_cols['total_spent_amount'])
            pnl_r_cell = worksheet.cell(row=2, column=summary_cols['pnl_r'])

            if pnl_r_cell.value is not None and total_spent_cell.value is not None:
                try:
                    pnl_r_value = float(pnl_r_cell.value)
                    total_spent_value = float(total_spent_cell.value)
                    if pnl_r_value > total_spent_value:
                        pnl_r_cell.fill = gold_fill
                    else:
                        pnl_r_cell.fill = red_fill
                except (ValueError, TypeError):
                    pass

        # Format dexscreener column width
        if 'dexscreener' in transaction_cols:
            dexscreener_column_letter = get_column_letter(transaction_cols['dexscreener'])
            worksheet.column_dimensions[dexscreener_column_letter].width = 20

        # Format transaction data (starting from transaction_start_row + 1)
        required_cols = ['delta_sol', 'delta_percentage']
        if all(col in transaction_cols for col in required_cols):
            for row in worksheet.iter_rows(min_row=transaction_start_row + 1):
                if len(row) == 0:
                    continue

                try:
                    # Get cells for this row
                    delta_percentage_cell = row[transaction_cols['delta_percentage'] - 1]
                    delta_sol_cell = row[transaction_cols['delta_sol'] - 1]

                    # Format delta_percentage and delta_sol
                    if delta_percentage_cell.value is not None:
                        percentage_value = float(delta_percentage_cell.value)

                        if percentage_value == -100:
                            delta_percentage_cell.fill = brown_fill
                            delta_sol_cell.fill = red_fill
                        elif percentage_value > 0:
                            delta_percentage_cell.fill = green_fill
                            delta_sol_cell.fill = green_fill
                        elif percentage_value < 0:
                            delta_percentage_cell.fill = red_fill
                            delta_sol_cell.fill = red_fill

                    # Format dexscreener hyperlink
                    if 'dexscreener' in transaction_cols:
                        dexscreener_cell = row[transaction_cols['dexscreener'] - 1]
                        if dexscreener_cell.value:
                            original_url = str(dexscreener_cell.value)
                            dexscreener_cell.value = "View Dexscreener"
                            dexscreener_cell.hyperlink = original_url
                            dexscreener_cell.font = Font(color="0000FF", underline="single")

                    # Additional formatting for number_buys if available
                    if 'number_buys' in transaction_cols:
                        buys_cell = row[transaction_cols['number_buys'] - 1]
                        if buys_cell.value is not None and float(buys_cell.value) > 3:
                            buys_cell.fill = yellow_fill

                except (ValueError, TypeError, IndexError):
                    continue  # Skip problematic rows

        print("✅ Combined sheet formatting applied successfully!")

    """-----------------------------FETCHING SOL TRANSFERS DATA-----------------------------------------------------"""

    def fetch_sol_transfers_data(self):
        """Fetch SOL transfers data from Dune"""
        sol_transfers_query = QueryBase(query_id= self.SOL_TRANSFER_QUERY_ID, params=self.parameters_transfer)

        self.sol_transfers_df = self.dune.run_query_dataframe(sol_transfers_query, performance='')
        self.sol_transfers_df.columns = [col.lower() for col in self.sol_transfers_df.columns]

        # Calculate EUR values
        if not self.sol_transfers_df.empty:
            self.sol_transfers_df['sol_amount_eur'] = self.sol_transfers_df['sol_amount'] * self.solana_eur_price
            self.sol_transfers_df['sol_eur_price'] = self.solana_eur_price



    def get_sol_transfers_from_db(self, days_back=None):
        """Get SOL transfers data from database"""
        query = """
            SELECT 
                wallet_id,
                sol_eur_price,
                block_month,
                from_owner,
                to_owner,
                sol_amount,
                sol_amount_eur,
                transaction_label,
                solscan_link
            FROM sol_transfers 
            WHERE wallet_id = ?
        """

        params = [self.wallet_id]
        if days_back:
            query += f" AND created_at >= datetime('now', '-{days_back} days')"

        query += " ORDER BY created_at DESC"

        return pd.read_sql_query(query, self.conn, params=params)

    def save_sol_transfers_to_database(self):
        """Save SOL transfers data to SQLite database using chunked insertion"""
        try:
            if not hasattr(self, 'sol_transfers_df') or self.sol_transfers_df.empty:
                print("❌ No SOL transfers data to save")
                return False

            print(f"Preparing to save {len(self.sol_transfers_df)} SOL transfers to database...")

            # Add wallet_id column to the dataframe
            self.sol_transfers_df['wallet_id'] = self.wallet_id
            print(f"Added wallet_id {self.wallet_id} to SOL transfers dataframe")

            # Debug: Check dataframe before saving
            print("SOL Transfers DataFrame info before saving:")
            print(f"Shape: {self.sol_transfers_df.shape}")
            print(f"Columns ({len(self.sol_transfers_df.columns)}): {list(self.sol_transfers_df.columns)}")

            # Calculate appropriate chunk size
            num_columns = len(self.sol_transfers_df.columns)
            max_chunk_size = max(1, 900 // num_columns)  # Leave buffer for SQLite's 999 limit
            chunk_size = min(max_chunk_size, 100)  # Cap at 100 for performance

            print(f"Using chunk size of {chunk_size} rows (with {num_columns} columns)")

            # Split DataFrame into chunks and save each chunk
            total_rows = len(self.sol_transfers_df)
            chunks = [self.sol_transfers_df[i:i + chunk_size] for i in range(0, total_rows, chunk_size)]

            print(f"Saving {len(chunks)} chunks...")

            for i, chunk in enumerate(chunks, 1):
                try:
                    chunk.to_sql(
                        name='sol_transfers',
                        con=self.conn,
                        if_exists='append',
                        index=False,
                        method='multi'
                    )
                    # print(f"✅ Saved SOL transfers chunk {i}/{len(chunks)} ({len(chunk)} rows)")

                except Exception as chunk_error:
                    print(f"❌ Error saving SOL transfers chunk {i}: {chunk_error}")
                    # Try with method=None for this chunk
                    try:
                        chunk.to_sql(
                            name='sol_transfers',
                            con=self.conn,
                            if_exists='append',
                            index=False,
                            method=None
                        )
                        print(f"✅ Saved SOL transfers chunk {i}/{len(chunks)} using fallback method")
                    except Exception as fallback_error:
                        print(f"❌ Fallback also failed for SOL transfers chunk {i}: {fallback_error}")
                        raise fallback_error

            self.conn.commit()
            print(f"✅ All SOL transfers data successfully saved to database for wallet ID: {self.wallet_id}")

            # Verify data was actually saved
            cursor = self.conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM sol_transfers WHERE wallet_id = ?", (self.wallet_id,))
            count = cursor.fetchone()[0]
            print(f"✅ Verification: {count} SOL transfer records found in database for wallet ID: {self.wallet_id}")

            return True

        except Exception as e:
            print(f"❌ Error saving SOL transfers to database: {e}")
            print(f"Error type: {type(e).__name__}")
            import traceback
            print("Full traceback:")
            print(traceback.format_exc())
            self.conn.rollback()
            return False

    def generate_sol_transfers_summary_from_db(self, days_back=None):
        """Generate SOL transfers summary from database"""
        query = """
            SELECT 
                w.wallet_address,
                COUNT(*) as total_transactions,
                SUM(CASE WHEN st.transaction_label = 'Sent' THEN st.sol_amount ELSE 0 END) as total_sent_sol,
                SUM(CASE WHEN st.transaction_label = 'Received' THEN st.sol_amount ELSE 0 END) as total_received_sol,
                SUM(CASE WHEN st.transaction_label = 'Sent' THEN st.sol_amount_eur ELSE 0 END) as total_sent_eur,
                SUM(CASE WHEN st.transaction_label = 'Received' THEN st.sol_amount_eur ELSE 0 END) as total_received_eur,
                COUNT(CASE WHEN st.transaction_label = 'Sent' THEN 1 END) as sent_count,
                COUNT(CASE WHEN st.transaction_label = 'Received' THEN 1 END) as received_count
            FROM sol_transfers st
            JOIN wallets w ON st.wallet_id = w.id
            WHERE st.wallet_id = ?
        """

        params = [self.wallet_id]
        days_back=self.days_back

        if days_back:
            query += f" AND st.created_at >= datetime('now', '-{days_back} days')"

        query += " GROUP BY w.wallet_address"

        # return pd.read_sql_query(query, self.conn, params=params)
        summary_df = pd.read_sql_query(query, self.conn, params=params)

        if days_back:
            summary_df['time_period_days'] = days_back
        else:
            summary_df['time_period_days'] = days_back   #All time string before

        return summary_df



    def apply_sol_transfers_formatting(self, input_file_path=None, output_file_path=None):
        """Apply formatting specifically for SOL transfers Excel"""

        if input_file_path is None:
            input_file_path = self.output_file_path
        if output_file_path is None:
            output_file_path = input_file_path

        workbook = load_workbook(input_file_path)

        # Check if required sheets exist
        if 'SOL Transfers Summary' not in workbook.sheetnames or 'SOL Transfers' not in workbook.sheetnames:
            print("❌ Required SOL transfer sheets not found. Skipping formatting.")
            return workbook

        # Access sheets
        summary_sheet = workbook['SOL Transfers Summary']
        transfers_sheet = workbook['SOL Transfers']

        # Create combined sheet
        if 'SOL Transfers Report' in workbook.sheetnames:
            del workbook['SOL Transfers Report']

        combined_sheet = workbook.create_sheet('SOL Transfers Report', 0)

        # Copy summary data (excluding unwanted columns)
        current_row = 1
        for row in range(1, summary_sheet.max_row + 1):
            col_index = 1
            for col in range(1, summary_sheet.max_column + 1):
                header_value = summary_sheet.cell(row=1, column=col).value
                # Skip wallet_id and sol_eur_price columns
                if header_value and (
                        'wallet_id' in str(header_value).lower() or 'sol_eur_price' in str(header_value).lower()):
                    continue

                value = summary_sheet.cell(row=row, column=col).value
                combined_sheet.cell(row=current_row, column=col_index).value = value
                col_index += 1
            current_row += 1

        # Add spacing
        current_row += 1

        # Copy transfers data (excluding unwanted columns)
        transfers_start_row = current_row
        for row in range(1, transfers_sheet.max_row + 1):
            col_index = 1
            for col in range(1, transfers_sheet.max_column + 1):
                header_value = transfers_sheet.cell(row=1, column=col).value
                # Skip wallet_id and sol_eur_price columns
                if header_value and (
                        'wallet_id' in str(header_value).lower() or 'sol_eur_price' in str(header_value).lower()):
                    continue

                value = transfers_sheet.cell(row=row, column=col).value
                combined_sheet.cell(row=current_row, column=col_index).value = value
                col_index += 1
            current_row += 1

        # Apply SOL-specific formatting
        self.apply_sol_transfers_specific_formatting(combined_sheet, transfers_start_row)

        # Remove original sheets
        del workbook['SOL Transfers Summary']
        del workbook['SOL Transfers']

        workbook.save(output_file_path)
        print(f'✅ SOL transfers formatted and saved to {output_file_path}')
        return workbook

    def apply_sol_transfers_specific_formatting(self, worksheet, transfers_start_row):
        """Apply specific formatting for SOL transfers"""

        # Define colors
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # Apply basic formatting
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                cell.alignment = Alignment(horizontal="center", vertical="center")

                # Bold headers
                if cell.row == 1 or cell.row == transfers_start_row:
                    cell.font = Font(bold=True)

        # Adjust column widths
        for column_cells in worksheet.columns:
            length = max(len(str(cell.value or '')) for cell in column_cells)
            worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        # Find columns for SOL transfers
        transfer_cols = {}
        for cell in worksheet[transfers_start_row]:
            if cell.value:
                col_name = str(cell.value).lower()
                if 'transaction_label' in col_name:
                    transfer_cols['transaction_label'] = cell.column
                elif 'solscan_link' in col_name:
                    transfer_cols['solscan_link'] = cell.column
                elif 'sol_amount' in col_name and 'eur' not in col_name:
                    transfer_cols['sol_amount'] = cell.column

        # Format transfer data
        for row in worksheet.iter_rows(min_row=transfers_start_row + 1):
            if len(row) == 0:
                continue

            try:
                # Color code based on transaction type
                if 'transaction_label' in transfer_cols:
                    label_cell = row[transfer_cols['transaction_label'] - 1]
                    if label_cell.value == 'Received':
                        label_cell.fill = green_fill
                        if 'sol_amount' in transfer_cols:
                            row[transfer_cols['sol_amount'] - 1].fill = green_fill
                    elif label_cell.value == 'Sent':
                        label_cell.fill = red_fill
                        if 'sol_amount' in transfer_cols:
                            row[transfer_cols['sol_amount'] - 1].fill = red_fill

                # Format Solscan links
                if 'solscan_link' in transfer_cols:
                    link_cell = row[transfer_cols['solscan_link'] - 1]
                    if link_cell.value and 'solscan.io' in str(link_cell.value):
                        original_url = str(link_cell.value)
                        link_cell.value = "View on Solscan"
                        link_cell.hyperlink = original_url
                        link_cell.font = Font(color="0000FF", underline="single")

            except (ValueError, TypeError, IndexError):
                continue

        print("✅ SOL transfers specific formatting applied!")

    def save_sol_transfers_to_excel(self):
        """Save SOL transfers data to Excel (from fetched data)"""
        # Update filename for SOL transfers
        self.output_file_path = os.path.join(
            self.reports_folder,
            f"{self.wallet_address}_SOL_transfers.xlsx"
        )

        # Get summary from database
        sol_transfers_summary_df = self.generate_sol_transfers_summary_from_db()

        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            # Write SOL transfers summary
            if not sol_transfers_summary_df.empty:
                sol_transfers_summary_df.to_excel(writer, sheet_name='SOL Transfers Summary', index=False)

            # Write SOL transfers details using fetched DataFrame
            if hasattr(self, 'sol_transfers_df') and not self.sol_transfers_df.empty:
                self.sol_transfers_df.to_excel(writer, sheet_name='SOL Transfers', index=False)
        self.apply_sol_transfers_formatting()

    def save_sol_transfers_excel_from_db(self, days_back=None):
        """Save SOL transfers Excel report using only database data"""
        # Get SOL transfers data from database
        sol_transfers_df = self.get_sol_transfers_from_db(days_back=days_back)
        sol_transfers_summary_df = self.generate_sol_transfers_summary_from_db(days_back=days_back)

        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            # Write SOL transfers summary
            if not sol_transfers_summary_df.empty:
                sol_transfers_summary_df.to_excel(writer, sheet_name='SOL Transfers Summary', index=False)

            # Write SOL transfers details from database
            if not sol_transfers_df.empty:
                sol_transfers_df.to_excel(writer, sheet_name='SOL Transfers', index=False)
        self.apply_sol_transfers_formatting()



    def generate_sol_transfers_excel_from_db(self, days_back=None):
        """Generate Excel report with SOL transfers from database only"""
        print(f"Generating SOL transfers Excel from database for wallet: {self.wallet_address} (ID: {self.wallet_id})")

        # Check if we have SOL transfers data in the database
        existing_data = self.get_sol_transfers_from_db(days_back=days_back)

        if existing_data.empty:
            print("❌ No SOL transfers data found in database for this wallet.")
            print("💡 Run option 3 first to fetch SOL transfers from Dune.")
            return False

        print(f"📊 Found {len(existing_data)} SOL transfers in database")

        # Generate Excel with time period in filename for SOL transfers
        if days_back:
            self.output_file_path = os.path.join(
                self.reports_folder,
                f"{self.wallet_address}_SOL_transfers_{days_back}days.xlsx"
            )
        else:
            self.output_file_path = os.path.join(
                self.reports_folder,
                f"{self.wallet_address}_SOL_transfers_all_time.xlsx"
            )

        # Save to Excel using database data
        self.save_sol_transfers_excel_from_db(days_back=days_back)

        print(f"✅ SOL transfers Excel report generated: {self.output_file_path}")

        # Display summary
        summary = self.generate_sol_transfers_summary_from_db(days_back=days_back)
        if not summary.empty:
            print(f"\nSOL Transfers Summary ({'Last ' + str(days_back) + ' days' if days_back else 'All time'}):")
            print(summary.to_string(index=False))

        return True

    def run_sol_transfers_report(self):
        """Method to fetch SOL transfers from Dune and save to database"""
        print(f"Fetching SOL transfers for wallet: {self.wallet_address} (ID: {self.wallet_id})")

        # Fetch SOL transfers data from Dune
        print("Fetching SOL transfers data from Dune...")
        self.fetch_sol_transfers_data()

        # Save SOL transfers to database
        print("Saving SOL transfers data to database...")
        self.save_sol_transfers_to_database()

        # Generate Excel with SOL transfers only
        print("Saving SOL transfers to Excel...")
        self.save_sol_transfers_to_excel()

        print(f"SOL transfers report completed. Files saved:")
        print(f"- Excel: {self.output_file_path}")

        # Display SOL transfers summary
        summary = self.generate_sol_transfers_summary_from_db()
        if not summary.empty:
            print("\nSOL Transfers Summary:")
            print(summary.to_string(index=False))


    """---------------Wallet Transactions Data Processing and Saving---------------------"""

    def fetch_data(self):
        """Fetch transaction data from Dune"""
        transaction_query = QueryBase(query_id=self.TRANSACTION_QUERY_ID, params=self.parameters)

        # self.transaction_df = self.dune.run_query_dataframe(transaction_query, performance='')
        # self.transaction_df.columns = [col.lower() for col in self.transaction_df.columns]
        try:
            self.transaction_df = self.dune.run_query_dataframe(transaction_query, performance='')
            self.transaction_df.columns = [col.lower() for col in self.transaction_df.columns]

            if self.transaction_df is not None:
                print(f"✅ Fetched {len(self.transaction_df)} rows from Dune")
                print(f"Columns: {list(self.transaction_df.columns)}")

                # Convert column names to lowercase
                self.transaction_df.columns = [col.lower() for col in self.transaction_df.columns]

                if len(self.transaction_df) > 0:
                    print("Sample data (first 2 rows):")
                    print(self.transaction_df.head(2).to_string())
                else:
                    print("⚠️ DataFrame is empty - no transactions found for this wallet/time period")
            else:
                print("❌ Failed to fetch data - transaction_df is None")

        except Exception as e:
            print(f"❌ Error fetching data from Dune: {e}")
            self.transaction_df = None

    def calculate_eur_values(self):
        """Calculate EUR values in bulk using vectorized operations"""
        if self.transaction_df is not None and not self.transaction_df.empty:
            # Bulk calculation using pandas vectorized operations (very fast)
            self.transaction_df['spent_amount_eur'] = self.transaction_df['spent_amount'] * self.solana_eur_price
            self.transaction_df['earned_amount_eur'] = self.transaction_df['earned_amount'] * self.solana_eur_price
            self.transaction_df['sol_eur_price'] = self.solana_eur_price  # Store the price used for calculation

            print(f"EUR calculations completed for {len(self.transaction_df)} records in bulk")



    def save_to_database(self):
        """Save the fetched transaction data to SQLite database using chunked insertion"""
        try:
            # Check if we have data to save
            if self.transaction_df is None:
                print("❌ No transaction data to save - transaction_df is None")
                return False

            if self.transaction_df.empty:
                print("❌ No transaction data to save - transaction_df is empty")
                return False

            print(f"Preparing to save {len(self.transaction_df)} transactions to database...")

            # Add wallet_id column to the dataframe
            self.transaction_df['wallet_id'] = self.wallet_id
            print(f"Added wallet_id {self.wallet_id} to dataframe")

            # Calculate EUR values in bulk before saving
            self.calculate_eur_values()

            # Debug: Check dataframe before saving
            print("DataFrame info before saving:")
            print(f"Shape: {self.transaction_df.shape}")
            print(f"Columns ({len(self.transaction_df.columns)}): {list(self.transaction_df.columns)}")

            # Calculate appropriate chunk size
            # SQLite limit is 999 variables, with ~17 columns, we can safely use chunks of 50 rows
            num_columns = len(self.transaction_df.columns)
            max_chunk_size = max(1, 900 // num_columns)  # Leave some buffer
            chunk_size = min(max_chunk_size, 100)  # Cap at 100 for performance

            print(f"Using chunk size of {chunk_size} rows (with {num_columns} columns)")

            # Split DataFrame into chunks and save each chunk
            total_rows = len(self.transaction_df)
            chunks = [self.transaction_df[i:i + chunk_size] for i in range(0, total_rows, chunk_size)]

            print(f"Saving {len(chunks)} chunks...")

            for i, chunk in enumerate(chunks, 1):
                try:
                    chunk.to_sql(
                        name='wallet_transactions',
                        con=self.conn,
                        if_exists='append',
                        index=False,
                        method='multi'
                    )
                    print(f"✅ Saved chunk {i}/{len(chunks)} ({len(chunk)} rows)")

                except Exception as chunk_error:
                    print(f"❌ Error saving chunk {i}: {chunk_error}")
                    # Try with method=None for this chunk
                    try:
                        chunk.to_sql(
                            name='wallet_transactions',
                            con=self.conn,
                            if_exists='append',
                            index=False,
                            method=None
                        )
                        print(f"✅ Saved chunk {i}/{len(chunks)} using fallback method")
                    except Exception as fallback_error:
                        print(f"❌ Fallback also failed for chunk {i}: {fallback_error}")
                        raise fallback_error

            self.conn.commit()
            print(f"✅ All transaction data successfully saved to database for wallet ID: {self.wallet_id}")

            # Verify data was actually saved
            cursor = self.conn.cursor()
            cursor.execute("SELECT COUNT(*) FROM wallet_transactions WHERE wallet_id = ?", (self.wallet_id,))
            count = cursor.fetchone()[0]
            print(f"✅ Verification: {count} records found in database for wallet ID: {self.wallet_id}")

            return True

        except Exception as e:
            print(f"❌ Error saving to database: {e}")
            print(f"Error type: {type(e).__name__}")
            import traceback
            print("Full traceback:")
            print(traceback.format_exc())
            self.conn.rollback()
            return False

    def generate_summary_from_db(self, days_back=None):
        """Generate comprehensive summary statistics from database data"""
        query = """
            SELECT 
                w.wallet_address AS wallet_id,
                COUNT(DISTINCT wt.token_symbol) AS number_of_tokens_traded,
                SUM(DISTINCT wt.spent_amount) AS total_spent_amount,
                SUM(DISTINCT wt.spent_amount_eur) AS total_spent_amount_eur,

                -- Actual profit calculation
                (SUM(CASE WHEN wt.delta_percentage > 0 THEN wt.delta_sol ELSE 0 END) - 
                 SUM(CASE WHEN wt.delta_percentage < 0 THEN wt.delta_sol ELSE 0 END) - 
                 SUM(DISTINCT wt.spent_amount)) AS actual_profit_sol,

                (SUM(CASE WHEN wt.delta_percentage > 0 THEN wt.earned_amount_eur - wt.spent_amount_eur ELSE 0 END) - 
                 SUM(CASE WHEN wt.delta_percentage < 0 THEN ABS(wt.earned_amount_eur - wt.spent_amount_eur) ELSE 0 END)) AS actual_profit_eur,

                -- Profits and losses
                SUM(CASE WHEN wt.delta_percentage > 0 THEN wt.delta_sol ELSE 0 END) AS pnl_realized_profits_sol,
                SUM(CASE WHEN wt.delta_percentage < 0 THEN wt.delta_sol ELSE 0 END) AS pnl_realized_losses_sol,
                SUM(CASE WHEN wt.delta_percentage > 0 THEN wt.earned_amount_eur - wt.spent_amount_eur ELSE 0 END) AS pnl_realized_profits_eur,
                SUM(CASE WHEN wt.delta_percentage < 0 THEN wt.earned_amount_eur - wt.spent_amount_eur ELSE 0 END) AS pnl_realized_losses_eur

            FROM wallet_transactions wt
            JOIN wallets w ON wt.wallet_id = w.id
            WHERE wt.wallet_id = ?
        """

        params = [self.wallet_id]
        days_back= self.days_back
        if days_back:
            query += f" AND wt.created_at >= datetime('now', '-{days_back} days')"

        query += " GROUP BY w.wallet_address, w.wallet_name"

        summary_df = pd.read_sql_query(query, self.conn, params=params)

        # Add time period info to the result
        if days_back:
            summary_df['time_period_days'] = days_back
        else:
            summary_df['time_period_days'] = days_back   #All time string before

        return summary_df







    def update_wallet_info(self, wallet_name=None, description=None):
        """Update wallet information"""
        cursor = self.conn.cursor()

        if wallet_name:
            cursor.execute('''
                UPDATE wallets SET wallet_name = ?, updated_at = CURRENT_TIMESTAMP 
                WHERE id = ?
            ''', (wallet_name, self.wallet_id))

        if description:
            cursor.execute('''
                UPDATE wallets SET description = ?, updated_at = CURRENT_TIMESTAMP 
                WHERE id = ?
            ''', (description, self.wallet_id))

        self.conn.commit()
        print(f"Wallet information updated for wallet ID: {self.wallet_id}")

    def reorder_columns(self):
        """Reorder DataFrame columns for better presentation"""
        desired_order = [
            'wallet_id', 'sol_eur_price', 'token_symbol', 'time_traded',
            'incoming', 'outcome', 'delta_token', 'spent_amount', 'spent_amount_eur',
            'earned_amount', 'earned_amount_eur', 'number_buys', 'number_sells',
            'delta_sol', 'delta_percentage', 'dexscreener', 'block_time'
        ]

        # Only reorder columns that exist
        existing_cols = [col for col in desired_order if col in self.transaction_df.columns]
        remaining_cols = [col for col in self.transaction_df.columns if col not in existing_cols]

        self.transaction_df = self.transaction_df[existing_cols + remaining_cols]



    def generate_excel_from_db(self, days_back=None):
        """Generate Excel report from existing database data without fetching from Dune"""
        print(f"Generating Excel report from database for wallet: {self.wallet_address} (ID: {self.wallet_id})")

        # Check if we have data in the database
        existing_data = self.get_wallet_transactions_from_db(days_back=days_back)

        if existing_data.empty:
            print("❌ No data found in database for this wallet.")
            print("💡 Run run_report() first to fetch data from Dune, or check if wallet has transactions.")
            return False

        print(f"📊 Found {len(existing_data)} transactions in database")

        # Generate Excel with time period in filename
        if days_back:
            self.output_file_path = os.path.join(
                self.reports_folder,
                f"{self.wallet_address}_{days_back}days.xlsx"
            )

        # Save to Excel using database data
        self.save_to_excel_from_db(days_back=days_back)

        print(f"✅ Excel report generated: {self.output_file_path}")

        # Display summary
        summary = self.generate_summary_from_db(days_back=days_back)
        if not summary.empty:
            print(f"\nWallet Summary ({'Last ' + str(days_back) + ' days' if days_back else 'All time'}):")
            print(summary.to_string(index=False))

        return True

    def get_wallet_transactions_from_db(self, days_back=None):
        """Retrieve wallet transaction data from database"""
        if days_back is None:
            days_back = self.days_back

        query = """
            SELECT wt.*
            FROM wallet_transactions wt
            JOIN wallets w ON wt.wallet_id = w.id
            WHERE wt.wallet_id = ?
        """
        params = [self.wallet_id]

        if days_back:
            query += " AND wt.created_at >= datetime('now', '-{} days')".format(days_back)

        query += " ORDER BY wt.block_time DESC"

        return pd.read_sql_query(query, self.conn, params=params)


    def save_to_excel(self):
        """Save data to Excel - now includes both transactions and generated summary with advanced formatting"""
        # Get summary from database
        summary_df = self.generate_summary_from_db()

        # Reorder columns before saving to Excel
        self.reorder_columns()

        print("Creating Excel file...")
        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            # Write summary first
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Write transaction details
            self.transaction_df.to_excel(writer, sheet_name='Transactions', index=False)

        # Apply advanced formatting after saving
        print("Applying advanced formatting...")
        self.combine_and_format_sheets_integrated()

    def save_to_excel_from_db(self, days_back=None):
        """Save Excel report using only database data with advanced formatting"""
        # Get summary from database
        summary_df = self.generate_summary_from_db(days_back=days_back)

        # Get transaction data directly from database
        transactions_df = self.get_wallet_transactions_from_db(days_back=days_back).drop(columns=['created_at'],
                                                                                         errors='ignore')

        print("Creating Excel file from database data...")
        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            # Write summary first
            if not summary_df.empty:
                summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # Write transaction details from database
            if not transactions_df.empty:
                transactions_df.to_excel(writer, sheet_name='Transactions', index=False)

        # Apply advanced formatting after saving
        print("Applying advanced formatting...")
        self.combine_and_format_sheets_integrated()


    def close_connection(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()

    def __del__(self):
        """Ensure database connection is closed"""
        self.close_connection()



if __name__ == "__main__":
    wallet_address = input("Enter Your Solana Wallet Address: ")
    days_back_input = input("Enter days back for fetching data (or press Enter for 15 days default): ")
    days_back = int(days_back_input) if days_back_input.strip() else 15

    report = SOLReport(wallet_address, days_back)  # Pass days_back parameter

    print("\nChoose an option:")
    print("1. Fetch wallet transactions data from Dune")
    print("2. Generate wallet transactions Excel from existing database data")
    print("3. Fetch SOL transfers data from Dune")
    print("4. Generate SOL transfers Excel from existing database data")

    choice = input("Enter your choice (1/2/3/4): ").strip()

    if choice == "1":
        # Fetch wallet transactions from Dune
        print("Fetching wallet transactions from Dune...")
        report.fetch_data()
        # print("Calculating EUR values")
        # report.calculate_eur_values()
        print("Saving wallet transactions to database...")
        report.save_to_database()
        print("Saving wallet transactions to Excel...")
        report.save_to_excel()

    elif choice == "2":
        # Generate wallet transactions Excel from database
        # days = input("Generate wallet transactions report for how many days back? (or press Enter for all): ").strip()
        days = days_back
        days_back = int(days) if days else None
        report.get_wallet_transactions_from_db(days_back=days_back)
        report.generate_summary_from_db(days_back=days_back)
        report.generate_excel_from_db(days_back=days_back)

    elif choice == "3":
        # Fetch SOL transfers from Dune
        report.run_sol_transfers_report()

    elif choice == "4":
        # Generate SOL transfers Excel from database
        days = days_back
        days_back = int(days) if days else None
        print("Fetching SOL transfers from database...")
        report.generate_sol_transfers_summary_from_db(days_back=days_back)
        print("Generating SOL transfers Excel from database...")
        report.generate_sol_transfers_excel_from_db(days_back=days_back)

    else:
        print("Invalid choice. Please select 1, 2, 3, or 4.")

    report.close_connection()


